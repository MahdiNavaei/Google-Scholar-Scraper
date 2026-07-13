import tempfile
import unittest
from pathlib import Path
import csv

from openpyxl import load_workbook

from google_scholar_scraper.exporters import EXPORT_COLUMNS, save_to_csv, save_to_excel
from google_scholar_scraper.models import Article


class ExporterTests(unittest.TestCase):
    def test_export_receives_normalized_article_rows(self) -> None:
        article = Article(
            title="Deep Learning",
            authors="A Author",
            link="https://example.edu/paper",
            relevance_score=87.5,
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "articles.xlsx"

            save_to_excel([article], path)

            worksheet = load_workbook(path).active

        rows = list(worksheet.iter_rows(values_only=True))
        self.assertEqual(list(rows[0]), ["Title", "Authors", "Link", "Relevance Score"])
        self.assertEqual(rows[1][0], "Deep Learning")
        self.assertEqual(rows[1][1], "A Author")
        self.assertEqual(rows[1][2], "https://example.edu/paper")
        self.assertEqual(rows[1][3], 87.5)

    def test_export_uses_blank_relevance_score_when_unranked(self) -> None:
        article = Article(title="Deep Learning", authors="A Author", link="https://example.edu/paper")
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "articles.xlsx"

            save_to_excel([article], path)

            worksheet = load_workbook(path).active

        rows = list(worksheet.iter_rows(values_only=True))
        self.assertEqual(list(rows[0]), ["Title", "Authors", "Link", "Relevance Score"])
        self.assertIsNone(rows[1][3])

    def test_csv_export_preserves_unicode_and_escapes_structured_text(self) -> None:
        articles = [
            Article(
                title='یادگیری عمیق, "پزشکی"\nنسخه دوم',
                authors='Author, "A"',
                link="https://example.edu/paper?x=1,y=2",
                relevance_score=91.2,
            ),
            Article(title="Unranked", authors="", link=""),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "articles.csv"

            save_to_csv(articles, path)

            raw = path.read_bytes()
            self.assertTrue(raw.startswith(b"\xef\xbb\xbf"))
            with path.open(newline="", encoding="utf-8-sig") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(list(rows[0].keys()), EXPORT_COLUMNS)
        self.assertEqual(rows[0]["Title"], 'یادگیری عمیق, "پزشکی"\nنسخه دوم')
        self.assertEqual(rows[0]["Authors"], 'Author, "A"')
        self.assertEqual(rows[0]["Link"], "https://example.edu/paper?x=1,y=2")
        self.assertEqual(rows[0]["Relevance Score"], "91.2")
        self.assertEqual(rows[1]["Relevance Score"], "")


if __name__ == "__main__":
    unittest.main()
